"""
add_fx_data.py
──────────────────────────────────────────────────────────────────────────────
Appends three new sheets to multi_asset_universe.xlsx:

  fx_prices        — monthly adjusted-close prices for FX ETFs
  fx_returns       — monthly simple returns derived from prices
  fx_carry_signals — foreign 3M interbank rate minus USD 3M interbank rate

FX universe: FXE, FXY, FXB, FXA, FXC, FXF (core liquid currency ETFs)

Run from repo root:
    python scripts/add_fx_data.py
──────────────────────────────────────────────────────────────────────────────
"""
from __future__ import annotations

import io
import time
from pathlib import Path

import numpy as np
import pandas as pd
import requests
import yfinance as yf
from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parent.parent
WORKBOOK  = REPO_ROOT / "data" / "raw" / "multi_asset_universe.xlsx"

FX_ETF_MAP = {
    "FXE": "Euro",
    "FXY": "JapaneseYen",
    "FXB": "BritishPound",
    "FXA": "AustralianDollar",
    "FXC": "CanadianDollar",
    "FXF": "SwissFranc",
}

# FRED 3M interbank rates (IR3TIB01 series — OECD/BIS, annualised %)
FX_CARRY_SERIES = {
    "FXE": "IR3TIB01EZM156N",   # Euro area
    "FXY": "IR3TIB01JPM156N",   # Japan
    "FXB": "IR3TIB01GBM156N",   # United Kingdom
    "FXA": "IR3TIB01AUM156N",   # Australia
    "FXC": "IR3TIB01CAM156N",   # Canada
    "FXF": "IR3TIB01CHM156N",   # Switzerland
    "_USD": "IR3TIB01USM156N",  # United States (reference)
}

PULL_START = "2000-01-01"
FRED_BASE  = "https://fred.stlouisfed.org/graph/fredgraph.csv"


# ── helpers ────────────────────────────────────────────────────────────────────

def _pull_fx_prices(tickers: list[str]) -> pd.DataFrame:
    print(f"  yfinance: {tickers} … ", end="", flush=True)
    raw = yf.download(tickers, start=PULL_START, auto_adjust=True, progress=False)
    prices = raw["Close"].resample("ME").last()
    prices.index = prices.index.to_period("M").to_timestamp("M")
    prices.index.name = "date"
    print("done")
    return prices.sort_index()[tickers]


def _pull_fred(series_id: str, label: str) -> pd.Series:
    url = f"{FRED_BASE}?id={series_id}&observation_start={PULL_START}"
    for attempt in range(3):
        try:
            resp = requests.get(url, timeout=30)
            resp.raise_for_status()
            df = pd.read_csv(
                io.StringIO(resp.text),
                parse_dates=["observation_date"],
                index_col="observation_date",
            )
            df.columns = [label]
            df = df.replace(".", np.nan).astype(float)
            df.index = pd.to_datetime(df.index)
            monthly = df.resample("ME").last()
            monthly.index = monthly.index.to_period("M").to_timestamp("M")
            monthly.index.name = "date"
            n = df.dropna().shape[0]
            print(f"  FRED {series_id}: {n} obs")
            return monthly[label]
        except Exception as exc:
            if attempt < 2:
                time.sleep(2)
            else:
                raise RuntimeError(f"FRED {series_id} failed: {exc}") from exc


def _write_sheet(wb, sheet_name: str, df: pd.DataFrame) -> None:
    """Write (or overwrite) a sheet in an open openpyxl workbook."""
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    # Header row
    ws.cell(row=1, column=1, value="date")
    for c_idx, col in enumerate(df.columns, start=2):
        ws.cell(row=1, column=c_idx, value=col)

    # Data rows
    for r_idx, (idx, row) in enumerate(df.iterrows(), start=2):
        ws.cell(row=r_idx, column=1, value=idx.strftime("%Y-%m-%d"))
        for c_idx, val in enumerate(row, start=2):
            ws.cell(row=r_idx, column=c_idx, value=None if pd.isna(val) else float(val))

    print(f"  Sheet '{sheet_name}' written: {df.shape[0]} rows × {df.shape[1]} cols")


# ── main ───────────────────────────────────────────────────────────────────────

def main() -> None:
    print("━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
    print("FX DATA EXPANSION  →  multi_asset_universe.xlsx")
    print("━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")

    tickers = list(FX_ETF_MAP.keys())

    # ── 1. FX ETF prices ──────────────────────────────────────────────────────
    print("\n[1/3] FX ETF prices …")
    fx_prices = _pull_fx_prices(tickers)
    for tk in tickers:
        first = fx_prices[tk].dropna().index
        print(f"  {tk}: first obs {first[0].date() if len(first) else 'N/A'}")

    # ── 2. FX ETF returns ─────────────────────────────────────────────────────
    print("\n[2/3] FX ETF returns (pct_change) …")
    fx_returns = fx_prices.pct_change().iloc[1:]
    fx_returns.index.name = "date"

    # ── 3. FX carry signals ───────────────────────────────────────────────────
    print("\n[3/3] FX carry signals from FRED (3M interbank rate differentials) …")
    rates: dict[str, pd.Series] = {}
    for key, sid in FX_CARRY_SERIES.items():
        rates[key] = _pull_fred(sid, key)
        time.sleep(0.3)

    usd_rate = rates["_USD"]
    fx_carry = pd.DataFrame({
        etf: rates[etf] - usd_rate
        for etf in tickers
        if etf in rates
    })
    fx_carry.index.name = "date"
    # Restrict to period where ETF prices exist
    fx_carry = fx_carry.loc[fx_carry.index.isin(fx_prices.index)]

    # ── 4. Append sheets ──────────────────────────────────────────────────────
    print(f"\nAppending sheets to {WORKBOOK.name} …")
    wb = load_workbook(WORKBOOK)
    _write_sheet(wb, "fx_prices",        fx_prices)
    _write_sheet(wb, "fx_returns",       fx_returns)
    _write_sheet(wb, "fx_carry_signals", fx_carry)
    wb.save(WORKBOOK)

    print(f"\nExisting sheets: {wb.sheetnames}")
    print("\n━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
    print("DONE")
    print("━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")


if __name__ == "__main__":
    main()
