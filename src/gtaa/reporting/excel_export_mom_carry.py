"""Excel exporter for the three-sleeve GTAA Momentum + Carry strategy.

Exports 16 sheets to an .xlsx workbook using openpyxl via pd.ExcelWriter.
Each sheet is populated from the dict returned by run_gtaa_mom_carry_backtest()
and the GTAAMomCarryData object.
"""
from __future__ import annotations

import sys
from datetime import date
from pathlib import Path

import numpy as np
import pandas as pd


# ── Sheet name constants (order matters — drives the ReadMe table) ─────────────

SHEET_DESCRIPTIONS = [
    ("ReadMe",                         "Metadata and index of all sheets in this workbook"),
    ("Data Audit",                     "Asset coverage, return start dates, missing-value count per asset"),
    ("Gross Returns",                  "Monthly returns and growth of $1 for each sleeve and the combined portfolio"),
    ("Portfolio Statistics",           "Annualized return, volatility, IR, avg/max drawdown, turnover"),
    ("Raw Factor Weights",             "Summary statistics for all three sleeve raw weight panels (zero-sum)"),
    ("Equity Mom Raw Weights",         "Cross-sectional rank-standardized weights — Equity sleeve (zero-sum)"),
    ("Commodity Mom Raw Weights",      "Cross-sectional rank-standardized weights — Commodity sleeve (zero-sum)"),
    ("FI Carry Raw Weights",           "Cross-sectional rank-standardized weights — FI sleeve (zero-sum)"),
    ("CrossAsset Mom 1% Vol Weights",  "Equity + Commodity momentum sleeves scaled to 1% ex-ante vol"),
    ("FI Carry 1% Vol Weights",        "Fixed-Income Carry sleeve scaled to 1% ex-ante vol"),
    ("Final 1% Vol Portfolio Weights", "Equal-weight combination of all 3 sleeves, rescaled to 1%"),
    ("Covariances",                    "Final 36-month annualized population covariance matrix (ddof=0, ×12)"),
    ("Volatilities",                   "Per-asset annualized volatility from the final covariance diagonal"),
    ("Correlations",                   "Correlation matrix derived from the covariance matrix (cov_ij / vol_i / vol_j)"),
    ("Turnover",                       "Monthly one-way funded-portfolio turnover series"),
    ("QA Checks",                      "All validation checks with Pass/Fail status"),
]


# ── Helpers ────────────────────────────────────────────────────────────────────

def _bold_header(ws, row: int, n_cols: int | None = None) -> None:
    """Make the first cell (or n_cols cells) of the given row bold."""
    try:
        from openpyxl.styles import Font
        row_cells = list(ws.iter_rows(min_row=row, max_row=row))[0]
        limit = len(row_cells) if n_cols is None else min(n_cols, len(row_cells))
        for cell in row_cells[:limit]:
            cell.font = Font(bold=True)
    except Exception:
        pass  # graceful degradation if openpyxl internals unavailable


def _freeze_top_row(ws) -> None:
    try:
        ws.freeze_panes = "A2"
    except Exception:
        pass


def _write_df(ws, df: pd.DataFrame, start_row: int = 1, include_index: bool = True) -> int:
    """Write a DataFrame to a worksheet starting at start_row.

    Returns the row index of the next available row after the data.
    """
    from openpyxl.styles import Font

    if include_index:
        headers = [df.index.name or "Date"] + list(df.columns)
    else:
        headers = list(df.columns)

    # Header row
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col_idx, value=str(h) if h is not None else "")
        cell.font = Font(bold=True)

    # Data rows
    for r_offset, (idx, row) in enumerate(df.iterrows(), start=1):
        row_num = start_row + r_offset
        if include_index:
            idx_val = idx.date() if hasattr(idx, "date") else idx
            ws.cell(row=row_num, column=1, value=idx_val)
            for c_offset, val in enumerate(row, start=2):
                ws.cell(row=row_num, column=c_offset, value=_to_python(val))
        else:
            for c_offset, val in enumerate(row, start=1):
                ws.cell(row=row_num, column=c_offset, value=_to_python(val))

    return start_row + len(df) + 1


def _to_python(val):
    """Convert numpy scalar to Python native type."""
    if isinstance(val, (np.integer,)):
        return int(val)
    if isinstance(val, (np.floating,)):
        return float(val)
    if isinstance(val, (np.bool_,)):
        return bool(val)
    if isinstance(val, pd.Timestamp):
        return val.date()
    return val


def _write_square_matrix(ws, df: pd.DataFrame, start_row: int = 1) -> int:
    """Write a square labeled matrix (row label + column headers)."""
    from openpyxl.styles import Font

    # Column header row
    ws.cell(row=start_row, column=1, value="").font = Font(bold=True)
    for c_idx, col in enumerate(df.columns, start=2):
        cell = ws.cell(row=start_row, column=c_idx, value=str(col))
        cell.font = Font(bold=True)

    # Data rows with row labels
    for r_offset, (row_label, row) in enumerate(df.iterrows(), start=1):
        row_num = start_row + r_offset
        cell = ws.cell(row=row_num, column=1, value=str(row_label))
        cell.font = Font(bold=True)
        for c_offset, val in enumerate(row, start=2):
            ws.cell(row=row_num, column=c_offset, value=_to_python(val))

    return start_row + len(df) + 1


# ── Sheet writers ──────────────────────────────────────────────────────────────

def _write_readme(wb, result: dict, data, config: dict) -> None:
    ws = wb["ReadMe"]

    from openpyxl.styles import Font

    # ── Metadata block ────────────────────────────────────────────────────────
    final_rets = result["final_returns"]
    start_dt = final_rets.index[0].date() if len(final_rets) else "N/A"
    end_dt   = final_rets.index[-1].date() if len(final_rets) else "N/A"

    meta = [
        ("Strategy",        "Three-Sleeve GTAA: Cross-Asset Momentum FMP + Fixed-Income Carry FMP"),
        ("Backtest period",  f"{start_dt} — {end_dt}"),
        ("Sleeves",          "F1 Equity Momentum · F2 Commodity Momentum · F3 FI Carry"),
        ("Data sources",     "Yahoo Finance (ETF prices) · FRED (yield proxies)"),
        ("Python version",   sys.version.split()[0]),
        ("Date generated",   str(date.today())),
    ]

    row = 1
    for label, value in meta:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=value)
        row += 1

    row += 1  # blank separator

    # ── Sheet index table ─────────────────────────────────────────────────────
    ws.cell(row=row, column=1, value="Sheet Name").font = Font(bold=True)
    ws.cell(row=row, column=2, value="Description").font = Font(bold=True)
    row += 1

    for sheet_name, desc in SHEET_DESCRIPTIONS:
        ws.cell(row=row, column=1, value=sheet_name)
        ws.cell(row=row, column=2, value=desc)
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 80


def _write_data_audit(wb, data) -> None:
    ws = wb["Data Audit"]

    from openpyxl.styles import Font

    sleeves = [
        ("Equity",      data.equity_returns),
        ("Commodity",   data.commodity_returns),
        ("Fixed-Income", data.fi_returns),
    ]

    headers = ["Asset", "Sleeve", "First Date", "Last Date", "Total Months", "Missing Count"]
    for c_idx, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c_idx, value=h).font = Font(bold=True)

    row = 2
    for sleeve_name, df in sleeves:
        for asset in df.columns:
            col = df[asset].dropna()
            first_dt = col.index[0].date() if len(col) else None
            last_dt  = col.index[-1].date() if len(col) else None
            total    = len(df[asset])
            missing  = int(df[asset].isna().sum())
            ws.cell(row=row, column=1, value=str(asset))
            ws.cell(row=row, column=2, value=sleeve_name)
            ws.cell(row=row, column=3, value=first_dt)
            ws.cell(row=row, column=4, value=last_dt)
            ws.cell(row=row, column=5, value=len(col))
            ws.cell(row=row, column=6, value=missing)
            row += 1

    _freeze_top_row(ws)


def _write_gross_returns(wb, result: dict) -> None:
    ws = wb["Gross Returns"]

    eq_ret  = result["equity_fmp_returns"]
    com_ret = result["commodity_fmp_returns"]
    fi_ret  = result["fi_fmp_returns"]
    fin_ret = result["final_returns"]

    # Align all to final_returns index
    common_idx = fin_ret.index
    eq_aligned  = eq_ret.reindex(common_idx)
    com_aligned = com_ret.reindex(common_idx)
    fi_aligned  = fi_ret.reindex(common_idx)

    # Growth of $1
    eq_growth  = (1 + eq_aligned).cumprod()
    com_growth = (1 + com_aligned).cumprod()
    fi_growth  = (1 + fi_aligned).cumprod()
    fin_growth = (1 + fin_ret).cumprod()

    # Peak and drawdown for final GTAA
    peak     = fin_growth.cummax()
    drawdown = (fin_growth - peak) / peak

    out = pd.DataFrame({
        "Date":                  common_idx,
        "Equity FMP Return":     eq_aligned.values,
        "Commodity FMP Return":  com_aligned.values,
        "FI Carry FMP Return":   fi_aligned.values,
        "Final GTAA Return":     fin_ret.values,
        "Equity FMP ($1)":       eq_growth.values,
        "Commodity FMP ($1)":    com_growth.values,
        "FI Carry FMP ($1)":     fi_growth.values,
        "Final GTAA ($1)":       fin_growth.values,
        "Final GTAA Peak":       peak.values,
        "Final GTAA Drawdown":   drawdown.values,
    })

    from openpyxl.styles import Font
    headers = list(out.columns)
    for c_idx, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c_idx, value=h).font = Font(bold=True)

    for r_offset, row_data in out.iterrows():
        row_num = r_offset + 2
        for c_idx, val in enumerate(row_data, start=1):
            ws.cell(row=row_num, column=c_idx, value=_to_python(val))

    _freeze_top_row(ws)


def _write_portfolio_statistics(wb, result: dict) -> None:
    ws = wb["Portfolio Statistics"]

    from openpyxl.styles import Font

    final_rets = result["final_returns"]
    start_dt   = str(final_rets.index[0].date()) if len(final_rets) else "N/A"
    end_dt     = str(final_rets.index[-1].date()) if len(final_rets) else "N/A"
    n_months   = len(final_rets)

    # Extra header rows
    extra = [
        ("Start Date",      start_dt),
        ("End Date",        end_dt),
        ("Number of Months", n_months),
    ]

    row = 1
    for label, value in extra:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=value)
        row += 1

    row += 1  # blank line

    stats = result["stats"]
    # Header
    ws.cell(row=row, column=1, value="Strategy").font = Font(bold=True)
    for c_idx, col in enumerate(stats.columns, start=2):
        ws.cell(row=row, column=c_idx, value=str(col)).font = Font(bold=True)
    row += 1

    for name, stat_row in stats.iterrows():
        ws.cell(row=row, column=1, value=str(name))
        for c_idx, val in enumerate(stat_row, start=2):
            ws.cell(row=row, column=c_idx, value=_to_python(val))
        row += 1

    ws.column_dimensions["A"].width = 30
    for col_letter in ["B", "C", "D", "E", "F", "G", "H"]:
        ws.column_dimensions[col_letter].width = 20


def _write_raw_factor_summary(wb, result: dict) -> None:
    """Combined summary of all three sleeve raw weight panels."""
    ws = wb["Raw Factor Weights"]
    from openpyxl.styles import Font

    row = 1
    sleeves = [
        ("Factor 1a — Equity Momentum",  result["equity_raw_weights"]),
        ("Factor 1b — Commodity Trend",  result["commodity_raw_weights"]),
        ("Factor 2  — Fixed-Income Carry", result["fi_raw_weights"]),
    ]

    for sleeve_label, wts in sleeves:
        ws.cell(row=row, column=1, value=sleeve_label).font = Font(bold=True)
        ws.cell(row=row + 1, column=1, value=f"Period: {wts.index[0].date()} — {wts.index[-1].date()}, {len(wts)} months, {len(wts.columns)} assets")
        row += 2

        # Per-asset summary statistics
        stats = pd.DataFrame({
            "Asset":   wts.columns,
            "Mean":    wts.mean().values,
            "Std":     wts.std().values,
            "Min":     wts.min().values,
            "Max":     wts.max().values,
            "% Positive": (wts > 0).mean().values,
        })
        for c_idx, col in enumerate(stats.columns, start=1):
            ws.cell(row=row, column=c_idx, value=col).font = Font(bold=True)
        row += 1
        for _, stat_row in stats.iterrows():
            for c_idx, val in enumerate(stat_row, start=1):
                ws.cell(row=row, column=c_idx, value=_to_python(val))
            row += 1
        row += 1  # gap between sleeves

    # Note pointing to individual sheets
    note_row = row + 1
    ws.cell(row=note_row, column=1, value="Full time-series detail in: Equity Mom Raw Weights · Commodity Mom Raw Weights · FI Carry Raw Weights").font = Font(italic=True)


def _write_raw_weights(wb, sheet_name: str, weights: pd.DataFrame) -> None:
    ws = wb[sheet_name]
    _write_df(ws, weights.round(8), start_row=1, include_index=True)
    _freeze_top_row(ws)


def _write_scaled_weights_two_section(wb, result: dict) -> None:
    """CrossAsset Mom 1% Vol Weights: equity section then commodity section."""
    ws = wb["CrossAsset Mom 1% Vol Weights"]

    from openpyxl.styles import Font

    eq_w  = result["equity_scaled_weights"]
    com_w = result["commodity_scaled_weights"]

    ws.cell(row=1, column=1, value="=== Equity Momentum Scaled Weights (1% Vol) ===").font = Font(bold=True)
    next_row = _write_df(ws, eq_w.round(8), start_row=2, include_index=True)

    gap_row = next_row + 2
    ws.cell(row=gap_row, column=1, value="=== Commodity Momentum Scaled Weights (1% Vol) ===").font = Font(bold=True)
    _write_df(ws, com_w.round(8), start_row=gap_row + 1, include_index=True)

    _freeze_top_row(ws)


def _write_fi_scaled_weights(wb, result: dict) -> None:
    ws = wb["FI Carry 1% Vol Weights"]
    _write_df(ws, result["fi_scaled_weights"].round(8), start_row=1, include_index=True)
    _freeze_top_row(ws)


def _write_final_weights(wb, result: dict) -> None:
    ws = wb["Final 1% Vol Portfolio Weights"]
    _write_df(ws, result["final_weights"].round(8), start_row=1, include_index=True)
    _freeze_top_row(ws)


def _write_covariances(wb, result: dict) -> None:
    ws = wb["Covariances"]
    _write_square_matrix(ws, result["final_cov"], start_row=1)
    _freeze_top_row(ws)


def _write_volatilities(wb, result: dict) -> None:
    ws = wb["Volatilities"]

    from openpyxl.styles import Font

    vols = result["final_vols"]
    ws.cell(row=1, column=1, value="Asset").font = Font(bold=True)
    ws.cell(row=1, column=2, value="Ann. Vol (%)").font = Font(bold=True)

    for r_idx, (asset, vol) in enumerate(vols.items(), start=2):
        ws.cell(row=r_idx, column=1, value=str(asset))
        ws.cell(row=r_idx, column=2, value=_to_python(vol * 100))

    _freeze_top_row(ws)


def _write_correlations(wb, result: dict) -> None:
    ws = wb["Correlations"]
    _write_square_matrix(ws, result["final_corr"], start_row=1)
    _freeze_top_row(ws)


def _write_turnover(wb, result: dict) -> None:
    ws = wb["Turnover"]

    from openpyxl.styles import Font

    to = result["final_turnover"]
    ws.cell(row=1, column=1, value="Date").font = Font(bold=True)
    ws.cell(row=1, column=2, value="One-Way Turnover").font = Font(bold=True)

    for r_idx, (dt, val) in enumerate(to.items(), start=2):
        ws.cell(row=r_idx, column=1, value=dt.date() if hasattr(dt, "date") else dt)
        ws.cell(row=r_idx, column=2, value=_to_python(val))

    # Summary rows
    summary_row = len(to) + 3
    ws.cell(row=summary_row, column=1, value="Avg Monthly Turnover").font = Font(bold=True)
    ws.cell(row=summary_row, column=2, value=_to_python(to.mean()))
    ws.cell(row=summary_row + 1, column=1, value="Ann. Turnover").font = Font(bold=True)
    ws.cell(row=summary_row + 1, column=2, value=_to_python(to.mean() * 12))

    _freeze_top_row(ws)


def _write_qa_checks(wb, result: dict, data, config: dict) -> None:
    from gtaa.analytics.qa_checks import run_gtaa_qa_checks
    ws = wb["QA Checks"]
    qa_df = run_gtaa_qa_checks(result, data, config)
    _write_df(ws, qa_df, start_row=1, include_index=False)
    _freeze_top_row(ws)


# ── Public API ─────────────────────────────────────────────────────────────────

def export_gtaa_mom_carry_excel(
    result: dict,
    data,           # GTAAMomCarryData
    config: dict,
    output_path: str | Path,
) -> None:
    """Export the full GTAA Mom-Carry backtest to a 16-sheet Excel workbook.

    Args:
        result:      Dict returned by run_gtaa_mom_carry_backtest(config).
        data:        GTAAMomCarryData object (also available as result["data"]).
        config:      Strategy config dict (parsed YAML).
        output_path: Destination .xlsx path.
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # All sheet names in order (all ≤ 31 chars for Excel compatibility)
    sheet_names = [name for name, _ in SHEET_DESCRIPTIONS]

    import openpyxl
    wb = openpyxl.Workbook()
    wb.active.title = sheet_names[0]
    for name in sheet_names[1:]:
        wb.create_sheet(title=name)

    # ── Populate each sheet ───────────────────────────────────────────────────
    _write_readme(wb, result, data, config)
    _write_data_audit(wb, data)
    _write_gross_returns(wb, result)
    _write_portfolio_statistics(wb, result)
    _write_raw_factor_summary(wb, result)                                        # new summary
    _write_raw_weights(wb, "Equity Mom Raw Weights",    result["equity_raw_weights"])
    _write_raw_weights(wb, "Commodity Mom Raw Weights", result["commodity_raw_weights"])
    _write_raw_weights(wb, "FI Carry Raw Weights",      result["fi_raw_weights"])
    _write_scaled_weights_two_section(wb, result)
    _write_fi_scaled_weights(wb, result)
    _write_final_weights(wb, result)
    _write_covariances(wb, result)
    _write_volatilities(wb, result)
    _write_correlations(wb, result)
    _write_turnover(wb, result)
    _write_qa_checks(wb, result, data, config)

    wb.save(str(output_path))
